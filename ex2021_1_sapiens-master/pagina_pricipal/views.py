from django.shortcuts import render, get_object_or_404, redirect
from .models import Event, Room, Program,  AddIcon, SelectIcon, User
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail


#Excel sheet
from openpyxl import Workbook, workbook
from django.http import HttpResponse
from datetime import datetime
from datetime import timedelta

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

#not in use at the moment

"""if AutorizedCpfs.objects.all() == True:
    lista_bd = AutorizedCpfs.objects.all()
    lista_cpf = []
    for item in lista_bd:
        lista_cpf.append(item.cpf)"""

#@login_required(login_url='/')
def menuView(request): 
    if not request.user.is_authenticated:
        return redirect('/') 

    """send_mail('Primeiro django test mail', 'Se você está recebendo esse email é porquê você conseguiu', 'marcos.teste083@gmail.com', ['marcos.otavio10@gmail.com',])
    if AutorizedCpfs.objects.all() == True:
        lista_bd = AutorizedCpfs.objects.all()
        lista_cpf = []
        for item in lista_bd:
            lista_cpf.append(item.cpf)

        if request.user.last_name not in lista_cpf: 

            
            if request.user.is_superuser:
                pass
            else:
                user_data = [request.user.username, request.user.last_name] 
                request.user.delete()
                return redirect('/cpf-invalido/')"""

    #Serve para exibir apenas o primeiro nome do usuário no header
    #Se for super user exibe o username
    # Repete Para a maioria das views   
    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username

    event = Event.objects.all()[0]
    event_name = event.slug
    
    #Quando estiver hosteando mais de um evento alterar linha abaixo
    #Provavelmente adicionar um filtro para o evento específico
    rooms = Room.objects.all()
    
    return render(request, 'hall.html', {'event':event, 'rooms':rooms, 'slug':event_name, "user_name":user_name})


#slug1 é o slug do evento, slug2 é o slug da sala selecionada
def daysView(request, slug1, slug2):

    if not request.user.is_authenticated:
        return redirect('/')

    
    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username
    
    event = Event.objects.all()[0]
    

    #sala específica que a pessoa entrou
    sala = get_object_or_404(Room, slug=slug2) 

    rooms = Room.objects.all()
    
    sala_days= [x for x in range(1, event.days+1)] #Para eventos genericos
    
    #Caso não haja necessidade de uma tela de dias (A sala possui apenas 1 dia)
    if sala.days == 1:
        return redirect(roomsGenericView, slug1, slug2, 'unico')

    #Para a sapiens. Posteriormente tornar automarizado. Serve para o loop do html
    sala_days = {1:'19/07/2021',2:'20/07/2021',3:'21/07/2021'}

    return render(request, 'days.html', { "sala_days":sala_days, "room":sala, 'event':event, 'rooms':rooms, "user_name":user_name})


#slug1 é o slug do evento, slug2 é o slug da sala selecionada, slug3 é o dia
def roomsGenericView(request, slug1, slug2, slug3):

    if not request.user.is_authenticated:
        return redirect('/')

    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username


    event = Event.objects.all()[0]
    
    sala = get_object_or_404(Room, slug=slug2)

    #Apenas atividades da sala específica
    #Se não for dia único ele pega as atividades específicas por dia
    #Se for dia único, ele pega apenas a atividade daquela sala
    if slug3 == 'unico':
        room_activities = Program.objects.filter(room=sala.id)
    else:

        room_activities = Program.objects.filter(room=sala.id, day =slug3)
    


    #dar uma olhada na linha abaixo depois (apenas rooms da sapiens)
    rooms = Room.objects.all()

    return render(request, 'rooms.html', { 'event':event, 'activities': room_activities, 'room':sala, 'rooms':rooms, "user_name":user_name})

def padletCardView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    
    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username

    rooms = Room.objects.all()
    event = Event.objects.all()[0]



    return render(request, "padlet_cards.html", {'event':event, 'rooms':rooms, 'slug1':slug1, "user_name":user_name})



#slug1 is the event, slug2 é a conta da padlet, slug3 é a edição da sapiens 
def padletView(request, slug1, slug2, slug3):
    if not request.user.is_authenticated:
        return redirect('/')


    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username
        
    rooms = Room.objects.all()
    event = Event.objects.all()[0] 
    
    return render(request, 'padlet.html', {'event':event, 'rooms':rooms, 'slug1':slug1,'slug2':slug2,'conta':slug2, "edicao":slug3, "user_name":user_name})


def desopilarView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username

    event = Event.objects.all()[0]
    rooms = Room.objects.all()  

    return render(request, 'desopilar.html', {'event': event, 'rooms':rooms, 'slug1':slug1, "user_name":user_name})

def programacaoView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username

    event = Event.objects.all()[0]
    rooms = Room.objects.all()  

    return render(request, 'programacao.html', {'event': event, 'rooms':rooms, 'slug1':slug1, "user_name":user_name})


def apoiadoresView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username

    event = Event.objects.all()[0]
    rooms = Room.objects.all() 


    return render(request, 'apoiadores.html', {'event': event, 'rooms':rooms, 'slug1':slug1, "user_name":user_name})

#feito específico para a história da sapiens
def historiaView(request, slug1):
    if not request.user.is_authenticated:
        return redirect('/')

    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username

    event = Event.objects.all()[0]
    rooms = Room.objects.all() 


    return render(request, 'historia.html', {'event': event, 'rooms':rooms, 'slug1':slug1, "user_name":user_name})



def get_excell_users(request):
    
    
    users_list =  User.objects.all()

    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename={event}-usuarios.xlsx'.format(
        event = Event.objects.all()[0],
    )
    
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Usuários"

    columns = ['Posição','Nome completo','Curso', 'Email']
    row_num = 1

    #This atribue value to the columns headers
    for col_number, column_title in enumerate(columns,1):
        cell = worksheet.cell(row = row_num, column=col_number)
        cell.value = column_title

    #Define the data for each cell in the row
    for user in users_list:
        row_num +=1

        row=[
            row_num-1,
            str(user.first_name),
            user.last_name,
            user.email,
        ]

        for col_number, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_number)
            cell.value = cell_value
            cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')

    #for fix the width of th column
    dimensions = [7,55, 20, 30]
    for col_num, width in enumerate(dimensions, 1):
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = width

    workbook.save(response)

    return response


def loginRedirect(request):

    return render(request, 'v1bem_vindo.html', {})



"""def teste1(request):

    if not request.user.is_authenticated:
        return redirect('/')

    if not request.user.is_superuser: 
        user_name = request.user.first_name.split()
        user_name = user_name[0]
    else:
        user_name = request.user.username

    event = Event.objects.all()[0]
    rooms = Room.objects.all() 


    return render(request, 'historia.html', {'event': event, 'rooms':rooms, "user_name":user_name})"""
    


